from openpyxl.utils import get_column_letter
from json import dumps
import pandas as pd
import os
import json
from django.http import HttpResponse
from django.conf import settings
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from pymongo import MongoClient
from bson.json_util import ObjectId
import re
import datetime
cli = MongoClient('localhost',27017)
print("connection success",cli)
db = cli.mongodbSchema

def importExcel(request):
    wb = load_workbook(filename=request)
    ws = wb.active
    my_list = []
    duplicate= []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        try:
            if my_dict['name']:
                if db.language.find({"code":my_dict['code']}).count() > 0:
                    duplicate.append(my_dict)
                else:
                    my_list.append(my_dict)
        except:
            pass
    if len(duplicate) > 0:
        response = {"data":duplicate,"code":200}
    else:
        data = db.language.insert_many(my_list)
        response = {"data": "import success", "code": 204}
    return response
def LanaguaugeExport():
    data = list(db.language.aggregate([{"$project": {"_id": 0, "name": 1,"code":1}}]))
    # js_data = dumps(data, indent=1)
    # with open('data.json', 'w') as file:
    #     file.write(js_data)
    # data = db.language.find()
    js_data = dumps(data, indent=1)
    with open('excel.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('excel.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported_json_data.xlsx')
    df.to_excel('exported_json_data.xlsx')


def ImportTranslation(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    master_code = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        try:
            if my_dict['translation_ID']:
                my_list.append(my_dict)
                # print(my_list)
        except:
            pass
    for i in my_list:
        if db.translations.find({"language":request['language'],"translation.trans_id":str(i['translation_ID'])}).count() > 0:
            duplicates.append(i)
        else:
            suffix = re.search(r'\d.*', str(i['translation_ID'])).group()
            master = db.master_code.find({"id":suffix }).count()
            if master <= 0:
                master_code.append(i['translation_ID'])
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 204}
        return response
    elif len(master_code) > 0:
        response = {"data": master_code, "code": 208}
        return response
    else:
        for i in my_list:
            if db.translations.find({"language":request['language']}).count() > 0:
                try:
                    if i['version']:
                        suffix = re.search(r'\d.*', str(i['translation_ID'])).group()
                        master_Id = list(db.master_code.find({"id":  suffix}))
                        db.translations.update_one({"language": request['language']},
                            {"$push": {"translation": {"text": i['text'],"versions": [
                            {"trans": i['translation'],"version": i['version'],"currentVersion": True}],
                               "master_id": master_Id[0]['id'],"trans_id": str(i['translation_ID'])}}})
                    else:
                        suffix = re.search(r'\d.*', str(i['translation_ID'])).group()
                        master_Id = list(db.master_code.find({"id": suffix}))
                        db.translations.update_one({"language": request['language']},
                                                   {"$push": {"translation": {"text": i['text'], "versions": [
                                                       {"trans": i['translation'], "version":1,
                                                        "currentVersion": True}],
                                                                              "master_id": master_Id[0]['id'],
                                                                              "trans_id": str(i['translation_ID'])}}})
                except:pass
            else:
                try:
                    if i['version']:
                        suffix = re.search(r'\d.*', str(i['translation_ID'])).group()
                        master_Id = list(db.master_code.find({"id":suffix }))
                        db.translations.insert_one(
                            {"language": request['language'], "translation": [{"text": i['text'],
                                        "versions": [
                                                     {"trans": i['translation'],
                                                             "version": i['version'],
                                                                "currentVersion": True}],
                                                                    "master_id": master_Id[0]['id'],
                                                                    "trans_id":  str(i['translation_ID'])}]})
                    else:
                        suffix = re.search(r'\d.*', str(i['translation_ID'])).group()
                        master_Id = list(db.master_code.find({"id": suffix }))
                        db.translations.insert_one(
                            {"language": request['language'], "translation": [{"text": i['text'],
                                                                               "versions": [
                                                                                   {"trans": i['translation'],
                                                                                    "version": 1,
                                                                                    "currentVersion": True}],
                                                                               "master_id": master_Id[0]['id'],
                                                                               "trans_id": str(i['translation_ID'])}]})

                except:
                    pass
    response = {"data": "success", "code": 200}
    return response




def TranslationExport(request):
    data = list(db.translations.find({"language": request}))
    # print(data)
    translations = []
    for i in data[0]['translation']:
        text = i['text']
        trans_id =i['trans_id']
        for j in i['versions']:
            if j['currentVersion'] == True:
                version = j['version']
                trans = j['trans']
            else:
                pass
        row_data = {"Text": text, "Version": version, "Translation Id": trans_id, "Translation": trans}
        translations.append(row_data)

    js_data = dumps(translations, indent=1)
    with open('translation.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('translation.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-translation.xlsx')

def ImportMasterCode(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        # print(my_dict)
        try:
            data = list(db.master_code.find({"id":str(my_dict['id'])}))
            if len(data) > 0:
                if my_dict['id']:
                    duplicates.append(my_dict)
                else:
                    pass
            else:
                if my_dict['id']:
                    myDict = {"id":str(my_dict['id']),"text":my_dict['text']}
                    my_list.append(myDict)
                else:
                    pass
        except:
            pass
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 200}
        return response
    else :
        data = db.master_code.insert_many(my_list)
        response = {"data": "import success", "code": 204}
    return response


def MasterCodeExport():
    data = list(db.master_code.find({},{"_id":0,"type":0}))
    # print(data)
    js_data = dumps(data, indent=1)
    with open('mastercode.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('mastercode.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-code.xlsx')

def ImportRequirement(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    master_codes = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
 
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        # print(my_dict)

        try:
            try:
                country = my_dict['Country'].split(',')
                for i in country:

                    if db.country.find({"code":i.strip()}).count() > 0:
                        pass
                    else:
                        print(i)
                        response = {"data":i, "code": 204}
                        return response
            except:
                pass
            # req = db.additional_requirement.find({"value":str(my_dict['Value'])}).count()
            req = 0
            if req > 0:
                duplicates.append(my_dict)
            else:
                try:
                    my_dict['Requirement']
                    my_list.append(my_dict)
                except KeyError:
                    pass
            if my_dict['Type'] == 'Translation-ID':
                master_code = db.master_code.find({"id":str(my_dict['Value']).strip()}).count()
                if master_code > 0:
                    pass
                else:
                    master_codes.append(my_dict['Value'])
            else:
                pass
        except KeyError:
            pass
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 200}
        return response
    # elif len(master_codes) > 0:
    #     response = {"data": master_codes, "code": 201}
    #     return response
    else:
        for i in my_list:
            country = []
            order = db.additional_requirement.find({}).count() + 1
            static = False
            if i['Order'] > 0:
                order = i['Order']
            if i['Static'] == 'Yes':
                static = True
            value = ""
            if i['Type'] == 'Translation-ID':
                if i['Value'] !=None:
                    value = i['Value']
            else:
                value = i['Value']
            try:
                country = i['Country'].split(',')
            except:
                country = []
            if isinstance(country, list):
               pass
            else:
                country =[]
            db.additional_requirement.update_many(
                {"order": {"$gte": int(i['Order'])}},
                {"$inc": {"order": 1}},
            )
            data = db.additional_requirement.insert_one({"requirement":i['Requirement'],"type":i['Type'],
                                                     "value":str(value),"country":country,"static":static,"order":order})
        response = {"data": "sucess", "code": 208}
        return response



        # print(my_dict)
        # data = db.additional_requirement.insert_one({"requirement":my_dict['requirement'],})

def RequirementExport(request):
    data = list(db.additional_requirement.find({}).sort("order",1))
    requirements = []
    for i in data:
        if i['type'] == 'Translation-ID':
            name = ''
            static = "No"
            if i['static']:
                static = 'Yes'
            if i['country'] == '':
                pass
            else:
                try:
                    for j in i['country']:
                        name += j + ','
                except:
                    name = ''
            text = i['value']
            if db.master_code.find({"id": str(i['value'])}).count() > 0:
                value = list(db.master_code.find({"id": str(i['value'])}))
                text = i['value'] + "-" + value[0]['text']
            dict_values = {"Order":i['order'],"Requirement": i['requirement'], "Type": i['type'],
                           "Value":text, "Country": name,"Static requirement":static}
            requirements.append(dict_values)
        elif i['type'] == 'Asset':
            name = ''
            static = "No"
            if i['static']:
                static = 'Yes'
            if i['country'] == '':
                pass
            else:
                try:
                    country_name = db.country.find({"code": {"$in": i['country']}})
                    for j in i['country']:
                        name += j + ','
                except:
                    name = ''
            image = db.myapp_asset.find({'id': i['value']})
            dict_values = {"Order":i['order'],"Requirement": i['requirement'], "Type": i['type'],
                           "Value": "image",
                            "Country":  name,"Static requirement":static}
            requirements.append(dict_values)
        else:
            name = ''
            static = "No"
            if i['static']:
                static = 'Yes'
            if i['country'] == '':
                pass
            else:
                try:
                    country_name = db.country.find({"code": {"$in": i['country']}})
                    for j in i['country']:
                        name += j+ ','
                except:
                    name = ''
            dict_values = {"Order":i['order'],"Requirement": i['requirement'], "Type": i['type'], "Value": i['value'],
                           "Country":  name,"Static requirement":static}
            requirements.append(dict_values)

    js_data = dumps(requirements, indent=1)
    with open('requirements.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('requirements.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-requirements.xlsx')

def CountryExport():
    countries = []
    data = list(db.country.find({},{"_id":0}))
    for i in data:
        name = ''
        language_name = db.language.find({"code":{"$in":i['language']}})
        for j in language_name:
            name += j['name']+','
        dicts = {'Code':i['code'],'Name':i['name'],'Language':name}
        countries.append(dicts)
    js_data = dumps(countries, indent=1)
    with open('country.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('country.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-country.xlsx')
def ImportCountry(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicate = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        try:
            if my_dict['Code']:
                if db.country.find({"code": str(my_dict['Code']).strip()}).count() > 0:
                    duplicate.append(my_dict)
                else:
                    my_list.append(my_dict)
        except:
            pass
        try:
            languages = my_dict['Language'].split(',')
            for j in languages:
                if db.language.find({"code":j.strip()}).count() > 0:
                    pass
                else:
                    response = {"data":j,"code":202}
                    return response
        except KeyError:
            pass
    if len(duplicate) > 0:
        response = {"data": duplicate, "code": 200}
    else:
        for i in my_list:
            language = []
            try:
                language = i['Language'].split(',')
            except:
                language = []
            data = db.country.insert_one({"code":i['Code'].strip(),"name":i['Name'].strip(),"language":language})
        response = {"data": "import success", "code": 204}
    return response

def FactoryExport():
    factory =[]
    data = list(db.factory.find({},{"_id":0}))
    for i in data:
        factory.append({"Factory Location":i['location']['label'],"Packed In Translation":i['packed_in']['label'],"Factory Address Translation":i['address']['label']})
    js_data = dumps(factory, indent=1)
    with open('factory.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('factory.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-factory.xlsx')

def BlendExport(request):
    blends = []
    data = db.blend.find({})
    for i in data:
        try:
            blend_name_loose = list(db.master_code.find({"id": i['blend_loose']}))
            blend_loose = blend_name_loose[0]['id'] + '-' + blend_name_loose[0]['text']
            blendLoose_code = blend_name_loose[0]['id']
        except:
            blend_loose = ''
            blendLoose_code = ''
        try:
            blend_name_teaBag = list(db.master_code.find({"id": i['blend_tea_bag']}))
            blend_tea_bag = blend_name_teaBag[0]['id'] + '-' + blend_name_teaBag[0]['text']
            blendTeaBag_code = blend_name_teaBag[0]['id']
        except:
            blend_tea_bag = ''
            blendTeaBag_code = ''
        try:
            ingredient_name_teaBag = list(db.master_code.find({"id": i['ingredient_tea_bag']}))
            ingredient_tea_bag = ingredient_name_teaBag[0]['id'] + '-' + ingredient_name_teaBag[0]['text']
            ingredientTeaBag_code = ingredient_name_teaBag[0]['id']
        except:
            ingredient_tea_bag = ''
            ingredientTeaBag_code = ''
        try:
            ingredient_name_loose = list(db.master_code.find({"id": i['ingredient_loose']}))
            ingredient_loose = ingredient_name_loose[0]['id'] + '-' + ingredient_name_loose[0]['text']
            ingredientLoose_code = ingredient_name_loose[0]['id']
        except:
            ingredient_loose = ''
            ingredientLoose_code = ''
        try:
            tea_origin_name = list(db.master_code.find({"id": i['tea_origin']}))
            tea_origin = tea_origin_name[0]['id'] + '-' + tea_origin_name[0]['text']
            tea_origin_code = tea_origin_name[0]['id']
        except:
            tea_origin = ''
            tea_origin_code = ''

        my_dict = {"name": i['name'], 'blend_loose': blend_loose, "blend_tea_bag": blend_tea_bag,
                   "ingredient_tea_bag": ingredient_tea_bag,
                   "ingredient_loose": ingredient_loose, "tea_origin": tea_origin,
                   "blendLoose_code": blendLoose_code,
                   "blendTeaBag_code": blendTeaBag_code, "tea_origin_code": tea_origin_code,
                   "ingredientLoose_code": ingredientLoose_code, "ingredientTeaBag_code": ingredientTeaBag_code,
                   "range": i['range']}
        blends.append(my_dict)
    js_data = dumps(blends, indent=1)
    with open('blends.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('blends.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-blends.xlsx')

def IngredientExport(request):
    ingredients = []
    data = list(db.ingredient.find())
    for i in data:
        try:
            ingredient_name = list(db.master_code.find({"id": i['tea_bag']}))
            tea_bag = ingredient_name[0]['id'] + '-' + ingredient_name[0]['text']
        except:
            tea_bag = ''
        try:
            ingredient_name = list(db.master_code.find({"id": i['loose_trans']}))
            loose_name = ingredient_name[0]['id'] + '-' + ingredient_name[0]['text']
        except:
            loose_name = ''
        my_dict = {"Ingredient Name": i['name'], 'Tea Bag': tea_bag, "Loose Tea": loose_name}
        ingredients.append(my_dict)
    js_data = dumps(ingredients, indent=1)
    with open('ingredients.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('ingredients.json', encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-ingredients.xlsx')

def LegalNameExport(request):
    legalName = []
    data = list(db.legal_name.find())
    for i in data:
        try:
            translations = list(db.master_code.find({"id": i['translation']}))
            translation = translations[0]['id'] + '-' + translations[0]['text']
        except:
            translation = ''
        my_dict = {"name": i['name'], "translation": translation,}
        legalName.append(my_dict)
    js_data = dumps(legalName, indent=1)
    with open('Legalname.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('Legalname.json', encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-Legalname_3.xlsx')


def ImportBlend(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    master_codes = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        # print(my_dict)
        key = {key.strip(): value for key, value in my_dict.items()}
        print(key)
        if 'BLEND NAME' in key.keys():
            if key['BLEND NAME'] is not None:
                val = list(key.values())
                try:
                    blend = db.blend.find({"name":key['BLEND NAME']}).count()
                    if blend > 0:
                        print(blend)
                        duplicates.append(key)
                    else:
                        # blendLoose = db.master_code.find({"id": str(val[2])}).count()
                        # blendTeabag = db.master_code.find({"id":str(val[1])}).count()
                        # teaOrigin = db.master_code.find({"id": str(val[5])}).count()
                        # ingredientLoose = db.master_code.find({"id": str(val[4])}).count()
                        # ingredientTeaBag = db.master_code.find({"id": str(val[3])}).count()
                        # if blendLoose > 0 and blendTeabag > 0 and teaOrigin > 0 and ingredientLoose > 0 and ingredientTeaBag > 0:
                        my_list.append(key)
                        # else:
                        #     master_codes.append(key)
                        #     print(key)
                        # my_list.append(my_dict)
                except:
                    pass
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 204}
        return response
    # elif len(master_codes) > 0:
    #     response = {"data": master_codes, "code": 208}
        return response
    elif len(my_list) > 0:
        for i in my_list:
            val = list(i.values())
            Range = ''
            try:
                if str(val[6]) == 'Black Tea' or str(val[6]) == 'Black tea':
                    Range = 'Black tea'
                else:
                    Range = 'Green tea'
            except:
                pass
            data =  db.blend.insert_one({"name":val[0],
                                     "blend_loose":str(val[2]),"blend_tea_bag":str(val[1]),
                                     "tea_origin":str(val[5]),"ingredient_loose":str(val[4]),
                                     "ingredient_tea_bag":str(val[3]),"range":Range})
        response = {"data": "import success", "code": 200}
        return response
    else:
        response = {"code": 300}
        return response

def ImportLegalName(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    master_codes = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        # print(list(my_dict.values()))
        if 'LEGAL NAME' in my_dict.keys():
            if my_dict['LEGAL NAME'] is not None:
                val = list(my_dict.values())
                # print(val[1])
            legal_name = db.legal_name.find({"name":str(my_dict['LEGAL NAME'])}).count()
            if legal_name > 0:
                duplicates.append(my_dict)
            else:
                legalName = db.master_code.find({"id":str(val[1])}).count()
                if legalName > 0:
                    my_list.append(my_dict)
                else:
                    master_codes.append(my_dict)
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 200}
        return response
    elif len(master_codes) > 0:
        response = {"data": master_codes, "code": 201}
        return response
    else:
        for i in my_list:
            val = list(i.values())
            data = db.legal_name.insert_one({"translation":str(val[1]), "name":val[0]})
        response = {"data": "import success", "code": 204}
        return response
        # return response

def CountryRequirementExport(request):
    data = list(db.additional_requirement.find({ "$or": [ { "country": "all" }, { "country": request.GET['country']}]}))
    requirements = []
    for i in data:
        if i['type'] == 'Translation-ID':
            text = ''
            if i['value'] != '':
                val = list(db.master_code.find({"id": i['value']}))
                text = str(i['value']) +"-"+ val[0]['text']

            dict_values = {"requirement": i['requirement'], "type": i['type'],
                           "value": text }
            requirements.append(dict_values)
        elif i['type'] == 'Asset':
            image = db.myapp_asset.find({'id': i['value']})
            dict_values = {"requirement": i['requirement'], "type": i['type'],
                           "value": "image",
                           }
            requirements.append(dict_values)
        else:
            dict_values = {"requirement": i['requirement'], "type": i['type'], "value": i['value'],
                           }
            requirements.append(dict_values)

    js_data = dumps(requirements, indent=1)
    with open('requirements.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('requirements.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/'+request.GET['country']+'-requirement.xlsx')
from .textTranslation import textTranslation
def MaterCodeTranslationExport(request):
    master_code = request.GET['master_id']
    data = textTranslation(master_code)
    translations = []
    for i in data:
        translations.append({"Translation ID":i['translation_ID'],"Translation":i['translation'],"Language Code":i['language']})
    js_data = dumps(translations, indent=1)
    with open('masterCode_translations.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('masterCode_translations.json', encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/masterCode_translations.xlsx')

def ImportBlendRequirement(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    master_codes = []
    not_present = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        # print(my_dict)

        try:
            try:
                countries_str = my_dict.get('Countries')
                if countries_str is not None and isinstance(countries_str, str):
                    countries = countries_str.split(',')
                    for i in countries:
                        if db.country.find({"code": i}).count() > 0:
                            pass
                        else:
                            print(i)
                            response = {"data": i, "code": 204}
                            return response
            except KeyError:
                pass
            # req = db.additional_requirement.find({"value":str(my_dict['Value'])}).count()
            req = 0
            if req > 0:
                duplicates.append(my_dict)
            else:
                try:
                    my_dict['Requirement']
                    my_list.append(my_dict)
                except KeyError:
                    pass
            if my_dict['Type'] == 'Translation-ID':
                master_code = db.master_code.find({"id":str(my_dict['Value'])}).count()
                if master_code > 0:
                    pass
                else:
                    master_codes.append(my_dict['Value'])
            else:
                pass
            try:
                name = my_dict['Blend Name']
                if db.blend.find({"name":name }).count() > 0:
                    pass
                    # print(list(db.blend.find({"name":name })))
                else:
                    not_present.append(my_dict)
            except KeyError:
                pass
        except KeyError:
            pass
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 200}
        return response
    elif len(master_codes) > 0:
        response = {"data": master_codes, "code": 201}
        return response
    else:
        for i in my_list:
            country = []
            try:
                country = i['Countries'].split(',')
            except:
                country = []
            blend_id = list(db.blend.find({"name":i['Blend Name']}))
            data = db.blend_requirements.insert_one({"requirement":i['Requirement'],"type":i['Type'],"blend":str(blend_id[0]['_id']),
                                                            "category":i['Category'],"value":str(i['Value']),"country":country})
        response = {"data": "sucess", "code": 208}
        return response
def BlendRequirementExport(request):

    data = list(db.blend_requirements.find({"blend":request.GET['id']}))

    requirements = []
    for i in data:
        if i['type'] == 'Translation-ID':
            name = ''
            if i['country'] == '':
                pass
            elif i['country'] == 'All':
                name = 'All'
            else:
                try:
                    country_name = db.country.find({"code": {"$in": i['country']}})
                    for j in country_name:
                        name += j['name'] + ','
                except:
                    name = ''
            text = list(db.master_code.find({"id": i['value']}))
            dict_values = {"requirement": i['requirement'], "type": i['type'],
                           "value": text[0]['id']+"-"+text[0]['text'],"category":i['category'],"country": name}
            requirements.append(dict_values)
        elif i['type'] == 'Asset':
            name = ''
            if i['country'] == '':
                pass
            elif i['country'] == 'All':
                name = 'All'
            else:
                try:
                    country_name = db.country.find({"code": {"$in": i['country']}})
                    for j in country_name:
                        name += j['name'] + ','
                except:
                    name = ''
            image = db.myapp_asset.find({'id': i['value']})
            dict_values = {"requirement": i['requirement'], "type": i['type'],
                           "value": "image","category":i['category'],
                            "country": name}
            requirements.append(dict_values)
        else:
            name = ''
            if i['country'] == '':
                pass
            elif i['country'] == 'All':
                name = 'All'
            else:
                try:
                    country_name = db.country.find({"code": {"$in": i['country']}})
                    for j in country_name:
                        name += j['name'] + ','
                except:
                    name = ''
            dict_values = {"requirement": i['requirement'], "type": i['type'], "value": i['value'],"category":i['category'],
                           "country": name}
            requirements.append(dict_values)

    js_data = dumps(requirements, indent=1)
    with open('blend-requirements.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('blend-requirements.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-blend-requirements.xlsx')

def ImportItemNo(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    blends = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        # print(my_dict)
        try:
            data = list(db.itemNo.find({"item":str(my_dict['Item No'])}))
            if len(data) > 0:
                if my_dict['id']:
                    duplicates.append(my_dict)
                else:
                    pass
            else:
                if my_dict['Item No']:
                    blend_list = my_dict['Blend Name'].split(',')
                    bln = []
                    for b in blend_list:
                        blend = list(db.blend.find({"name":b}))
                        if len(blend) > 0:
                            bln.append({"label":blend[0]['name'],"value":blend[0]['_id']})
                        else:
                            blends.append({"name":b})
                    myDict = {"item": str(my_dict['Item No']), "blend_name": bln,
                               "category": my_dict['Category'],"number_of_teabag":str(my_dict['Number of Teabag'])}
                    my_list.append(myDict)
                else:
                    pass
        except:
            pass
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 200}
        return response
    elif len(blends) > 0:
        response = {"data":blends,"code":201}
    else :
        data = db.itemNo.insert_many(my_list)
        response = {"data": "import success", "code": 204}
    return response
def ExportItemNo(request):
    data = db.itemNo.find({})
    item_no = []
    for i in data:
        blend = ''
        for x in i['blend_name']:
            blend = blend + x['label']+',\n'
        item_no.append(
            {"Item No": i['item'], "Blend Name": blend,"Category":i['category'],"Number of Teabag":i['noTeabag']})
    js_data = dumps(item_no, indent=1)
    with open('itemNo.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('itemNo.json', encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/itemNumber.xlsx')

def ImportRegistrationNo(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    country  = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value
        # print(my_dict)
        try:
            data = list(db.registrationNo.find({"item":str(my_dict['Item No']),"registration":my_dict['Registration No']}))
            if len(data) > 0:
                if my_dict['id']:
                    country.append(my_dict)
                else:
                    pass
            if db.country.find({"code":my_dict['Country']}).count() < 0:
                country.append(my_dict['Country'])
            else:
                if my_dict['Item No']:
                    myDict = {"item":str(my_dict['Item No']),"registration":my_dict['Registration No'],"country":my_dict['Country']}
                    my_list.append(myDict)
                else:
                    pass
        except:
            pass
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 200}
        return response
    elif len(country) > 0:
        response = {"data": country, "code": 201}
        return response
    else :
        data = db.registrationNo.insert_many(my_list)
        response = {"data": "import success", "code": 204}
    return response

def ExportRegistrationNo(request):
    data = db.registrationNo.find({"item":request})
    reg_no = []
    for i in data:
        reg_no.append(
            {"Item No": i['item'], "Registation No": i['registration'], "Country": i['country']})
    js_data = dumps(reg_no, indent=1)
    with open('RegNo.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('RegNo.json', encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/RegistrationNumber'+str(request)+'.xlsx')

def ExportReport(request):
    data = request
    myList = []
    for i in data:
        mongo_date = i['date']
        timestamp = mongo_date["$date"] / 1000
        dt = datetime.datetime.fromtimestamp(timestamp)
        formatted_date = dt.strftime("%m/%d/%Y")

        mydict = {
            "Item": i['item'],
            "Varient": i['varient'],
            "Product": i['Description'],
            "Status": i['status'],
            "User": i['user'],
            "User Group": i['user_group'],
            "Date": formatted_date
        }
        myList.append(mydict)

    js_data = dumps(myList, indent=1)

    with open('Report.json', 'w') as file:
        file.write(js_data)

    df = pd.read_json('Report.json', encoding='utf-8-sig')
    df['Date'] = pd.to_datetime(df['Date'], format='%m/%d/%Y').dt.strftime('%d/%m/%Y')
    df.to_excel('media/exceltemplates/Report.xlsx')
def ItemRequirementExport(request):
    print(request.GET['item'])
    data = list(db.item_no_requirements.find({"item":str(request.GET['item'])}))

    requirements = []
    for i in data:
        if i['type'] == 'Translation-ID':
            name = ''
            if i['country'] == '':
                pass
            elif i['country'] == 'All':
                name = 'All'
            else:
                try:
                    country_name = db.country.find({"code": {"$in": i['country']}})
                    for j in country_name:
                        name += j['name'] + ','
                except:
                    name = ''
            text = list(db.master_code.find({"id": i['value']}))
            dict_values = {"requirement": i['requirement'], "type": i['type'],
                           "value": text[0]['id']+"-"+text[0]['text'],"country": name}
            requirements.append(dict_values)
        elif i['type'] == 'Asset':
            name = ''
            if i['country'] == '':
                pass
            elif i['country'] == 'All':
                name = 'All'
            else:
                try:
                    country_name = db.country.find({"code": {"$in": i['country']}})
                    for j in country_name:
                        name += j['name'] + ','
                except:
                    name = ''
            image = db.myapp_asset.find({'id': i['value']})
            dict_values = {"requirement": i['requirement'], "type": i['type'],
                           "value": "image",
                            "country": name}
            requirements.append(dict_values)
        else:
            name = ''
            if i['country'] == '':
                pass
            elif i['country'] == 'All':
                name = 'All'
            else:
                try:
                    country_name = db.country.find({"code": {"$in": i['country']}})
                    for j in country_name:
                        name += j['name'] + ','
                except:
                    name = ''
            dict_values = {"requirement": i['requirement'], "type": i['type'], "value": i['value'],
                           "country": name}
            requirements.append(dict_values)

    js_data = dumps(requirements, indent=1)
    with open('item-requirements.json', 'w') as file:
        file.write(js_data)
    df = pd.read_json('item-requirements.json',encoding='utf-8-sig')
    df.to_excel('media/exceltemplates/exported-item-requirements.xlsx')

def ImportItemRequirement(request):
    wb = load_workbook(filename=request['files'])
    ws = wb.active
    my_list = []
    duplicates = []
    master_codes = []
    not_present = []
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))
    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[ws[column_letter + str(1)].value] = ws[column_letter + str(row)].value


        try:
            try:
                country = my_dict['Countries'].split(',')
                for i in country:
                    if db.country.find({"code":i}).count() > 0:
                        pass
                    else:
                        print(i)
                        response = {"data":i, "code": 204}
                        return response
            except KeyError:
                pass
            # req = db.item_no_requirements.find({"value":str(my_dict['Value'])}).count()
            req = 0
            if req > 0:
                duplicates.append(my_dict)
            else:
                print(my_dict)
                try:
                    my_dict['Requirement'].strip()
                    print(my_dict)
                    my_list.append(my_dict)
                except KeyError:
                    pass
            if my_dict['Type'] == 'Translation-ID':
                master_code = db.master_code.find({"id":str(my_dict['Value']).strip()}).count()
                if master_code > 0:
                    pass
                else:
                    master_codes.append(my_dict['Value'])
            else:
                pass
            try:
                item = str(my_dict['Item Number']).strip()
                if db.itemNo.find({"item":item }).count() > 0:
                    pass
                    # print(list(db.blend.find({"name":name })))
                else:
                    not_present.append(my_dict)
            except KeyError:
                pass
        except KeyError:
            pass
    # print(my_list)
    # print(not_present)
    # print(duplicates)
    if len(duplicates) > 0:
        response = {"data": duplicates, "code": 200}
        return response
    elif len(master_codes) > 0:
        response = {"data": master_codes, "code": 201}
        return response
    else:
        for i in my_list:
            country = []
            try:
                country = i['Countries'].split(',')
            except:
                country = []
            data = db.item_no_requirements.insert_one({"requirement":i['Requirement'].strip(),"type":i['Type'],"item":str(i['Item Number']).strip(),
                                                            "value":str(i['Value']).strip(),"country":country})
        response = {"data": "sucess", "code": 208}
        return response

